import os
import time
from pathlib import Path

import requests
from dotenv import load_dotenv
from openpyxl import load_workbook

load_dotenv()

class Planilha:
    def __init__(self, diretorios, api):
        self.diretorios = diretorios
        self.API = api

        caminho_honorarias = Path(os.getenv("CONFISSOES"))
        self.wb = load_workbook(caminho_honorarias)
        self.sheet_honorarias = self.wb.active

        caminho_3C = Path(os.getenv("BASE_3C"))
        self.wb_3C = load_workbook(caminho_3C)
        self.sheet_3c = self.wb_3C.active

    def iterar_base(self):
        for linha in range (2, self.sheet_honorarias.max_row+1):
            time.sleep(.5)
            nome = self.sheet_honorarias.cell(row=linha, column=1).value
            if nome is None:
                break
            
            processo = self.sheet_honorarias.cell(row=linha, column=2).value

            try:
                print("\n============================================================\n")
                print(f"Buscando os dados do(a) {nome}.....")
                caminho, arquivo = self.diretorios.pegar_arquivo(nome)
                novo_caminho, novo_arquivo = self.diretorios.renomear_arquivo(processo, caminho, arquivo)
                self.API.upload(processo, novo_caminho, novo_arquivo)

            except:
                try:
                    print("Buscando o dados no 3C...")
                    nome_alternativo = self.buscar_nome_na_base_3C(processo)
                    if nome_alternativo is not None:
                        caminho, arquivo = self.diretorios.pegar_arquivo(nome_alternativo)

                        if caminho is None:
                            print(f"\n❌ O processo '{processo}' não possui um diretório")
                            self.criar_relatorio(nome, processo, "DIRETÓRIO INEXISTENTE")
                            continue
                        elif arquivo is None:
                            print(f"\n❌ O processo '{processo}' não possui um documento de confissão")
                            self.criar_relatorio(nome, processo, "NÃO POSSUI CONFISSÃO")
                            continue
                        novo_caminho, novo_arquivo = self.diretorios.renomear_arquivo(processo, caminho, arquivo)
                        
                        response = self.API.upload(processo, novo_caminho, novo_arquivo)
                        match response:
                            case 1:
                                ...
                            case 2:
                                self.criar_relatorio(nome, processo, "DIRETÓRIO INEXISTENTE")
                            case 3:
                                self.criar_relatorio(nome, processo, "PROCESSO NÃO LOCALIZADO NO SISTEMA")
                            case 4 :
                                self.criar_relatorio(nome, processo, "ERRO DE API")
                    else:
                        raise
                except:
                    print(f"\n❌ O nome: {nome} não foi localizado em lugar algum")
                    self.criar_relatorio(nome, processo, "NOME NÃO LOCALIZADO")
                                    
    def buscar_nome_na_base_3C(self, processo):
        for linha in range (2, self.sheet_3c.max_row+1):
            temp_processo = self.sheet_3c.cell(row=linha, column=2).value

            if str(temp_processo) == str(processo):
                nome_do_reu = self.sheet_3c.cell(row=linha, column=1).value
                return nome_do_reu            
        return None

    def criar_relatorio(self, nome, processo, razao):
        relatorio = [nome, processo, razao]
        self.wb = load_workbook("planilhas\RELATORIO.xlsx")
        sheet = self.wb.active

        try:
            sheet.append(relatorio)
            print(f"✅ Relatório de '{nome}' pre-armazenados com sucesso.")
        
        except Exception as e:
            print(f"❌ Erro ao armazenar relatório.             Detalhers do erro: {e}")
    
    def salvar_relatorio(self):
        self.wb.save("planilhas\RELATORIO.xlsx")

class Diretorios:
    def __init__(self):
        self.rede = Path("P:\\")        
    
    def pegar_arquivo(self, nome_do_reu):
        diretorio_do_reu = Path(os.path.join(self.rede, nome_do_reu))
        if diretorio_do_reu.is_dir():
            for elemento in diretorio_do_reu.rglob('*'):
                if "CONFISSÃO" in elemento.name.upper() or "CONFISSAO" in elemento.name.upper():
                    print(f"\n✅ Arquivo do réu: '{nome_do_reu}' localizado!!")
                    return elemento.parent, elemento.name
            return diretorio_do_reu, None
        return None, None
        
    def renomear_arquivo(self, processo, caminho, arquivo):
        novo_nome = f"{processo}_CONFISSAO_DE_DIVIDA"
        caminho_original = os.path.join(caminho, arquivo)
        novo_caminho = os.path.join(caminho, novo_nome)

        try:
            Path(caminho_original).rename(novo_caminho)
            print(f"✅ CONFISSÃO RENOMEADA PARA -> {novo_caminho}")
            return novo_caminho, novo_nome

        except Exception as e:
            print(f"\n❌ Não foi possível renomear o arquivo CONFISSÃO.\nDetalhes do erro: {e}")

class API:
    def __init__(self):
        self.TOKEN = os.getenv("TOKEN")
        self.URL = os.getenv("URL")

        # Configurar cabeçalhos
        self.headers = {
            "Authorization": f"Token {self.TOKEN}",
            "Accept": "application/json"
        }
    
    def upload(self, processo, caminho, arquivo):        
        file_path = Path(caminho)
        if not file_path.exists():
            print(f"❌ Arquivo não encontrado: {file_path}")
            return 0
        
        print(f"🟠 Enviando {arquivo} para o processos {processo} ...")
        with open(file_path, "rb") as f:
            files = {
                "document" : (file_path.name, f, "application/pdf")
            }
            response = requests.post(self.URL, files=files, headers=self.headers, timeout=10)
            match response.status_code:
                case 200:
                    print(f"✅ Upload realizado com sucesso.             Status: {response.status_code}\n")
                    return 1
                case 400:
                    print(f"❌ Arquivo não encontrado: {caminho}.             Status: {response.status_code}")
                    return 2
                case 404:
                    print(f"❌ Processos {processo} não foi encontrado no sistema.             Status: {response.status_code}")
                    return 3
                case _ :
                    print(f"❌ Erro de upload desconhecido: \n\nprocesso: {processo}\ncaminho do arquivo: {os.path.join(caminho or '', arquivo or '')}\nStatus: {response}")
                    return 4
            
if __name__ == "__main__":
    diretorios = Diretorios()
    api = API()
    
    base = Planilha(diretorios, api)
    base.iterar_base()
    base.salvar_relatorio() #BUG -> Criou apenas o relatório do último caso